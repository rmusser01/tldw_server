import asyncio
import base64
import shutil
import threading
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient

from tldw_Server_API.app.api.v1.endpoints import files as files_endpoint
from tldw_Server_API.app.core.AuthNZ.User_DB_Handling import User, get_request_user
from tldw_Server_API.app.core.config import settings
from tldw_Server_API.app.core.DB_Management.Collections_DB import CollectionsDatabase
from tldw_Server_API.app.core.DB_Management.db_path_utils import DatabasePaths


pytestmark = pytest.mark.integration

BASE_OPTIONS = {"persist": True}


_FILES_SENSITIVE_MARKERS = (
    "files path leaked",
    "local-export-dir",
    "777",
)


class _LoggerStub:
    def __init__(self):
        self.errors: list[tuple[str, tuple[object, ...], dict[str, object]]] = []
        self.warnings: list[tuple[str, tuple[object, ...], dict[str, object]]] = []

    def error(self, message: str, *args: object, **kwargs: object) -> None:
        self.errors.append((message, args, kwargs))

    def warning(self, message: str, *args: object, **kwargs: object) -> None:
        self.warnings.append((message, args, kwargs))


class _UnresolvableBaseDir:
    def resolve(self, *, strict: bool = False):
        raise RuntimeError("files path leaked local-export-dir 777")


@pytest.fixture()
def client_with_user(monkeypatch):
    async def override_user():
        return User(id=321, username="tester", email=None, is_active=True)

    monkeypatch.setenv("MINIMAL_TEST_APP", "0")
    monkeypatch.setenv("ULTRA_MINIMAL_APP", "0")
    monkeypatch.setenv("AUTH_MODE", "single_user")
    monkeypatch.setenv("SINGLE_USER_FIXED_ID", "321")

    base_dir = Path.cwd() / "Databases" / "test_user_dbs_files"
    shutil.rmtree(base_dir, ignore_errors=True)
    base_dir.mkdir(parents=True, exist_ok=True)
    users_db_path = base_dir / "users.db"
    prev_base_dir = settings.get("USER_DB_BASE_DIR")
    settings.USER_DB_BASE_DIR = str(base_dir)
    monkeypatch.setenv("USER_DB_BASE_DIR", str(base_dir))
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{users_db_path}")

    app = None
    try:
        from importlib import import_module, reload
        from tldw_Server_API.app.core.AuthNZ.database import reset_db_pool
        from tldw_Server_API.app.core.AuthNZ.initialize import ensure_single_user_rbac_seed_if_needed
        from tldw_Server_API.app.core.DB_Management.Users_DB import reset_users_db

        # Ensure settings/pools pick up the test database and seed the single-user row at id=321.
        asyncio.run(reset_db_pool())
        asyncio.run(reset_users_db())
        asyncio.run(ensure_single_user_rbac_seed_if_needed())

        mod = import_module("tldw_Server_API.app.main")
        mod = reload(mod)
        app = mod.app
        app.dependency_overrides[get_request_user] = override_user
        with TestClient(app) as client:
            yield client
    finally:
        if app is not None:
            app.dependency_overrides.clear()
        if prev_base_dir is not None:
            settings.USER_DB_BASE_DIR = prev_base_dir
        else:
            try:
                del settings.USER_DB_BASE_DIR
            except AttributeError:
                pass


def test_resolve_export_path_sanitizes_base_dir_failure_log(monkeypatch):
    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(
        files_endpoint.DatabasePaths,
        "get_user_temp_outputs_dir",
        lambda _user_id: _UnresolvableBaseDir(),
    )

    with pytest.raises(HTTPException) as exc_info:
        files_endpoint._resolve_export_path_for_user(777, "export.md")

    assert exc_info.value.status_code == 500
    assert exc_info.value.detail == "storage_unavailable"
    assert logger_stub.errors == [("files: failed to resolve temp outputs base dir", (), {})]
    rendered = " ".join([logger_stub.errors[0][0], *(str(arg) for arg in logger_stub.errors[0][1])])
    for marker in _FILES_SENSITIVE_MARKERS:
        assert marker not in rendered


def test_clear_export_state_sanitizes_cleanup_failure_log(monkeypatch):
    class _FailingCollectionsDb:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def update_file_artifact_export(self, *args, **kwargs):
            raise RuntimeError("export cleanup leaked /private/files/export-state.db")

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(
        files_endpoint.CollectionsDatabase,
        "for_user",
        staticmethod(lambda user_id: _FailingCollectionsDb()),
    )

    files_endpoint._clear_export_state(
        user_id=777,
        file_id=123,
        row=SimpleNamespace(
            export_format="md",
            export_bytes=10,
            export_content_type="text/markdown",
            export_job_id=None,
            export_expires_at=None,
        ),
        consumed_at=None,
    )

    assert logger_stub.warnings == [("files.export: failed to clear export state", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export-state.db" not in rendered


@pytest.mark.asyncio
async def test_export_expired_file_delete_failure_log_is_sanitized(monkeypatch):
    class _FakeExportPath:
        def exists(self) -> bool:
            return True

        def unlink(self) -> None:
            raise RuntimeError("expired export unlink leaked /private/files/export.md")

    class _FakeCollectionsDb:
        def get_file_artifact(self, file_id: int):
            assert file_id == 123
            return SimpleNamespace(
                export_status="ready",
                export_storage_path="export.md",
                export_format="md",
                export_expires_at=(datetime.now(timezone.utc) - timedelta(minutes=1)).isoformat(),
            )

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", lambda user_id, path: _FakeExportPath())
    monkeypatch.setattr(files_endpoint, "_clear_export_state", lambda **kwargs: None)

    with pytest.raises(HTTPException) as exc_info:
        await files_endpoint.export_file_artifact(
            file_id=123,
            format="md",
            cdb=_FakeCollectionsDb(),
            current_user=SimpleNamespace(id=777),
        )

    assert exc_info.value.status_code == 404
    assert exc_info.value.detail == "export_expired"
    assert logger_stub.warnings == [("files.export: failed to delete expired export file", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


@pytest.mark.asyncio
async def test_export_background_file_delete_failure_log_is_sanitized(monkeypatch):
    class _FakeExportPath:
        name = "export.md"

        def __fspath__(self) -> str:
            return "/tmp/export.md"

        def exists(self) -> bool:
            return True

        def unlink(self) -> None:
            raise RuntimeError("export unlink leaked /private/files/export.md")

    class _FakeCollectionsDb:
        def get_file_artifact(self, file_id: int):
            assert file_id == 123
            return SimpleNamespace(
                export_status="ready",
                export_storage_path="export.md",
                export_format="md",
                export_content_type="text/markdown",
                export_bytes=10,
                export_job_id=None,
                export_expires_at=(datetime.now(timezone.utc) + timedelta(minutes=1)).isoformat(),
                export_consumed_at=None,
            )

        def consume_file_artifact_export(self, file_id: int, *, consumed_at: str) -> bool:
            assert file_id == 123
            assert consumed_at
            return True

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", lambda user_id, path: _FakeExportPath())
    monkeypatch.setattr(files_endpoint, "_clear_export_state", lambda **kwargs: None)

    response = await files_endpoint.export_file_artifact(
        file_id=123,
        format="md",
        cdb=_FakeCollectionsDb(),
        current_user=SimpleNamespace(id=777),
    )

    assert response.background is not None
    await response.background()

    assert logger_stub.warnings == [("files.export: failed to delete export file", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


@pytest.mark.asyncio
async def test_delete_file_artifact_invalid_export_path_log_is_sanitized(monkeypatch):
    class _FakeCollectionsDb:
        def get_file_artifact(self, file_id: int, *, include_deleted: bool = False):
            assert file_id == 123
            assert include_deleted is True
            return SimpleNamespace(export_storage_path="export.md")

        def delete_file_artifact(self, file_id: int, *, hard: bool = False) -> bool:
            assert file_id == 123
            assert hard is True
            return True

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)

    def _raise_invalid_path(user_id: int, path: str):
        raise HTTPException(status_code=400, detail="leaked invalid path /private/files/export.md")

    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", _raise_invalid_path)

    response = await files_endpoint.delete_file_artifact(
        file_id=123,
        hard=True,
        delete_file=True,
        cdb=_FakeCollectionsDb(),
        current_user=SimpleNamespace(id=777),
    )

    assert response.success is True
    assert response.file_deleted is False
    assert logger_stub.warnings == [("files.delete: invalid export path", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


@pytest.mark.asyncio
async def test_delete_file_artifact_delete_failure_log_is_sanitized(monkeypatch):
    class _FakeExportPath:
        def exists(self) -> bool:
            return True

        def unlink(self) -> None:
            raise RuntimeError("delete unlink leaked /private/files/export.md")

    class _FakeCollectionsDb:
        def get_file_artifact(self, file_id: int, *, include_deleted: bool = False):
            assert file_id == 123
            assert include_deleted is True
            return SimpleNamespace(export_storage_path="export.md")

        def delete_file_artifact(self, file_id: int, *, hard: bool = False) -> bool:
            assert file_id == 123
            assert hard is True
            return True

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", lambda user_id, path: _FakeExportPath())

    response = await files_endpoint.delete_file_artifact(
        file_id=123,
        hard=True,
        delete_file=True,
        cdb=_FakeCollectionsDb(),
        current_user=SimpleNamespace(id=777),
    )

    assert response.success is True
    assert response.file_deleted is False
    assert logger_stub.warnings == [("files.delete: failed to delete export file", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


@pytest.mark.asyncio
async def test_purge_file_artifacts_invalid_export_path_log_is_sanitized(monkeypatch):
    class _FakeCollectionsDb:
        def list_file_artifacts_for_purge(self, *, now_iso: str, soft_deleted_grace_days: int, include_retention: bool):
            assert now_iso
            assert soft_deleted_grace_days == 30
            assert include_retention is True
            return {123: "export.md"}

        def delete_file_artifacts_by_ids(self, file_ids: list[int]) -> int:
            assert file_ids == [123]
            return 1

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)

    def _raise_invalid_path(user_id: int, path: str):
        raise HTTPException(status_code=400, detail="leaked invalid path /private/files/export.md")

    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", _raise_invalid_path)

    response = await files_endpoint.purge_file_artifacts(
        payload=files_endpoint.FileArtifactsPurgeRequest(delete_files=True),
        cdb=_FakeCollectionsDb(),
        current_user=SimpleNamespace(id=777),
    )

    assert response.removed == 1
    assert response.files_deleted == 0
    assert logger_stub.warnings == [("files.purge: invalid export path", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


@pytest.mark.asyncio
async def test_purge_file_artifacts_delete_failure_log_is_sanitized(monkeypatch):
    class _FakeExportPath:
        def exists(self) -> bool:
            return True

        def unlink(self) -> None:
            raise RuntimeError("purge unlink leaked /private/files/export.md")

    class _FakeCollectionsDb:
        def list_file_artifacts_for_purge(self, *, now_iso: str, soft_deleted_grace_days: int, include_retention: bool):
            assert now_iso
            assert soft_deleted_grace_days == 30
            assert include_retention is True
            return {123: "export.md"}

        def delete_file_artifacts_by_ids(self, file_ids: list[int]) -> int:
            assert file_ids == [123]
            return 1

    logger_stub = _LoggerStub()
    monkeypatch.setattr(files_endpoint, "logger", logger_stub)
    monkeypatch.setattr(files_endpoint, "_resolve_export_path_for_user", lambda user_id, path: _FakeExportPath())

    response = await files_endpoint.purge_file_artifacts(
        payload=files_endpoint.FileArtifactsPurgeRequest(delete_files=True),
        cdb=_FakeCollectionsDb(),
        current_user=SimpleNamespace(id=777),
    )

    assert response.removed == 1
    assert response.files_deleted == 0
    assert logger_stub.warnings == [("files.purge: failed to delete export file", (), {})]
    rendered = " ".join([logger_stub.warnings[0][0], *(str(arg) for arg in logger_stub.warnings[0][1])])
    assert "123" not in rendered
    assert "/private/files/export.md" not in rendered


def test_create_and_export_markdown_table(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "md", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    data = response.json()
    artifact = data["artifact"]
    assert artifact["file_type"] == "markdown_table"
    assert artifact["export"]["status"] == "ready"
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "| Name | Score |" in download.text

    download_again = client_with_user.get(export_url)
    assert download_again.status_code == 409, download_again.text


def test_export_one_time_concurrent_downloads(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "md", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_url = response.json()["artifact"]["export"]["url"]
    assert export_url

    results = []
    errors = []
    lock = threading.Lock()
    barrier = threading.Barrier(2)

    def _download():
        try:
            barrier.wait(timeout=5)
            resp = client_with_user.get(export_url)
            with lock:
                results.append(resp.status_code)
        except Exception as exc:
            with lock:
                errors.append(str(exc))

    threads = [threading.Thread(target=_download) for _ in range(2)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    assert not errors, f"errors: {errors}"
    assert len(results) == 2
    assert results.count(200) == 1
    assert results.count(409) == 1


def test_get_file_artifact(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    artifact = fetch.json()["artifact"]
    assert artifact["file_id"] == artifact_id
    assert artifact["export"]["status"] == "none"


def test_export_not_ready_returns_404(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    export = client_with_user.get(f"/api/v1/files/{artifact_id}/export?format=md")
    assert export.status_code == 404, export.text


def test_async_export_returns_pending(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "md", "mode": "url", "async_mode": "async"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 202, response.text
    artifact = response.json()["artifact"]
    assert artifact["export"]["status"] == "pending"
    assert artifact["export"]["job_id"]


def test_delete_file_artifact_soft(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact_id = response.json()["artifact"]["file_id"]

    delete = client_with_user.delete(f"/api/v1/files/{artifact_id}")
    assert delete.status_code == 200, delete.text
    assert delete.json()["success"] is True

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 404, fetch.text


def test_create_and_export_xlsx(client_with_user):
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    payload = {
        "file_type": "xlsx",
        "title": "Roster",
        "payload": {
            "sheets": [
                {"name": "Sheet1", "columns": ["Name", "Score"], "rows": [["Ada", 95]]},
            ]
        },
        "export": {"format": "xlsx", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(download.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=2).value == "Score"
    assert ws.cell(row=2, column=1).value == "Ada"
    assert ws.cell(row=2, column=2).value == 95


def test_create_and_export_csv_table(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "Name,Score" in download.text
    assert "Ada,95" in download.text


def test_create_and_export_json_table(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "json", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code in (200, 201), response.text
    data = response.json()
    artifact = data["artifact"]
    assert artifact["export"]["status"] == "ready"
    assert artifact["export"]["format"] == "json"
    export_url = artifact["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    exported = download.json()
    expected_rows = [
        dict(zip(payload["payload"]["columns"], payload["payload"]["rows"][0], strict=True))
    ]
    assert exported == expected_rows


def test_create_requires_options(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_create_rejects_persist_false(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "options": {"persist": False},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_auto_async_uses_size_estimate(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Big",
        "payload": {"columns": ["Name", "Blob"], "rows": [["Ada", "x" * 5000]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "auto"},
        "options": {"persist": True, "max_bytes": 100},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 202, response.text
    artifact = response.json()["artifact"]
    assert artifact["export"]["status"] == "pending"


def test_create_warnings_on_duplicate_columns(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Dupes",
        "payload": {"columns": ["Name", "Name"], "rows": [["Ada", 95]]},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    warnings = response.json()["artifact"]["validation"]["warnings"]
    assert warnings
    assert warnings[0]["code"] == "duplicate_columns"


def test_inline_export_returns_content_b64(client_with_user, monkeypatch):
    monkeypatch.setenv("FILES_INLINE_MAX_BYTES", "1024")
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "md", "mode": "inline", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    export_info = artifact["export"]
    assert export_info["content_b64"]
    assert export_info["status"] == "none"
    assert export_info["url"] is None
    decoded = base64.b64decode(export_info["content_b64"]).decode("utf-8")
    assert "| Name |" in decoded
    artifact_id = artifact["file_id"]
    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    assert fetch.json()["artifact"]["export"]["status"] == "none"
    export = client_with_user.get(f"/api/v1/files/{artifact_id}/export?format=md")
    assert export.status_code == 409, export.text


def test_inline_export_falls_back_to_url(client_with_user, monkeypatch):
    monkeypatch.setenv("FILES_INLINE_MAX_BYTES", "10")
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada Lovelace"]]},
        "export": {"format": "md", "mode": "inline", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_info = response.json()["artifact"]["export"]
    assert export_info["url"]
    assert export_info["content_b64"] is None


def test_export_expired_clears_state(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "md", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    artifact = response.json()["artifact"]
    artifact_id = artifact["file_id"]
    export_url = artifact["export"]["url"]
    assert export_url

    cdb = CollectionsDatabase.for_user(user_id=321)
    row = cdb.get_file_artifact(artifact_id)
    past = (datetime.now(timezone.utc) - timedelta(seconds=1)).replace(microsecond=0).isoformat()
    cdb.update_file_artifact_export(
        artifact_id,
        export_status="ready",
        export_format=row.export_format,
        export_storage_path=row.export_storage_path,
        export_bytes=row.export_bytes,
        export_content_type=row.export_content_type,
        export_job_id=row.export_job_id,
        export_expires_at=past,
        export_consumed_at=None,
    )

    outputs_dir = DatabasePaths.get_user_temp_outputs_dir(321)
    if row.export_storage_path:
        export_path = outputs_dir / row.export_storage_path
        assert export_path.exists()
    expired = client_with_user.get(export_url)
    assert expired.status_code == 404, expired.text
    assert expired.json().get("detail") == "export_expired"
    if row.export_storage_path:
        assert not export_path.exists()

    fetch = client_with_user.get(f"/api/v1/files/{artifact_id}")
    assert fetch.status_code == 200, fetch.text
    assert fetch.json()["artifact"]["export"]["status"] == "none"


def test_invalid_export_format_returns_422(client_with_user):
    payload = {
        "file_type": "markdown_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_export_size_exceeded_returns_422(client_with_user):
    payload = {
        "file_type": "data_table",
        "title": "Roster",
        "payload": {"columns": ["Name"], "rows": [["Ada"]]},
        "export": {"format": "csv", "mode": "url", "async_mode": "sync"},
        "options": {"persist": True, "max_bytes": 1},
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 422, response.text


def test_create_and_export_html_table(client_with_user):
    payload = {
        "file_type": "html_table",
        "title": "Roster",
        "payload": {"columns": ["Name", "Score"], "rows": [["Ada", 95]]},
        "export": {"format": "html", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_url = response.json()["artifact"]["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "<table>" in download.text
    assert "<th>Name</th>" in download.text


def test_create_and_export_ical(client_with_user):
    pytest.importorskip("icalendar", reason="icalendar not installed")
    payload = {
        "file_type": "ical",
        "title": "Schedule",
        "payload": {
            "calendar": {
                "prodid": "-//tldw//files//EN",
                "version": "2.0",
                "timezone": "UTC",
                "events": [
                    {
                        "uid": "event-1",
                        "summary": "Kickoff",
                        "start": "2026-01-01T10:00:00",
                        "end": "2026-01-01T11:00:00",
                    }
                ],
            }
        },
        "export": {"format": "ics", "mode": "url", "async_mode": "sync"},
        "options": BASE_OPTIONS,
    }
    response = client_with_user.post("/api/v1/files/create", json=payload)
    assert response.status_code == 200, response.text
    export_url = response.json()["artifact"]["export"]["url"]
    assert export_url

    download = client_with_user.get(export_url)
    assert download.status_code == 200, download.text
    assert "BEGIN:VCALENDAR" in download.text
    assert "BEGIN:VEVENT" in download.text
